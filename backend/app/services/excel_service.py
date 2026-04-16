"""
Excel Export Service — PRD §4.1.5, §4.1.6, §4.1.10
openpyxl tabanlı .xlsx dosyası üretir.
"""
import io
from datetime import date, datetime
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Renkler ────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="3D3D3D")      # Koyu gri
SUBHEADER_FILL = PatternFill("solid", fgColor="5B8DEF")   # Accent mavi
ALT_ROW_FILL = PatternFill("solid", fgColor="2A2A2A")     # Alternatif satır
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
SUCCESS_FILL = PatternFill("solid", fgColor="2E7D32")
WARNING_FILL = PatternFill("solid", fgColor="ED6C02")
ERROR_FILL = PatternFill("solid", fgColor="C62828")

WHITE = "FFFFFF"
ACCENT = "5B8DEF"
TEXT_HEADER_COLOR = "E8E8E8"
TEXT_BODY_COLOR = "BDBDBD"

# ─── Kenarlık ──────────────────────────────────────────────────────────────
_thin = Side(style="thin", color="404040")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_font(bold: bool = True, color: str = WHITE) -> Font:
    return Font(name="Calibri", size=11, bold=bold, color=color)


def _body_font(size: int = 10, color: str = TEXT_BODY_COLOR) -> Font:
    return Font(name="Calibri", size=size, color=color)


def _set_header(ws, row: int, headers: List[str], col_start: int = 1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_width), max_width
        )


def _money(value: Any) -> str:
    if value is None:
        return ""
    try:
        return f"₺{float(value):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return str(value)


def _date(val: Any) -> Optional[date]:
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        try:
            return datetime.fromisoformat(val.replace("Z", "+00:00")).date()
        except Exception:
            pass
    return None


# ─── Finans / Transactions ──────────────────────────────────────────────────

def export_transactions_to_excel(transactions: List[Dict[str, Any]], agency_name: str = "") -> bytes:
    """
    Finansal işlemleri Excel'e döker — PRD §4.1.5.
    columns: Tarih | Açıklama | Kategori | Tür | Tutar | Durum
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Finansal İşlemler"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value=f"{agency_name} — Finansal İşlemler")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor="1A1A1A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Export date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    date_cell = ws.cell(row=2, column=1, value=f"Dışa Aktarım: {date.today().strftime('%d.%m.%Y')}")
    date_cell.font = _body_font(size=9, color="888888")
    date_cell.fill = PatternFill("solid", fgColor="1A1A1A")
    date_cell.alignment = Alignment(horizontal="left")

    # Headers
    _set_header(ws, 3, ["Tarih", "Açıklama", "Kategori", "Tür", "Tutar", "Durum"])

    # Data
    for i, tx in enumerate(transactions):
        row = i + 4
        tx_date = _date(tx.get("transaction_date")) or _date(tx.get("created_at"))

        # Status
        status = tx.get("status", "")
        status_color = ""
        if status == "confirmed" or status == "approved":
            status_color = "✓ Ödendi"
        elif status == "pending":
            status_color = "⏳ Bekliyor"
        elif status == "cancelled":
            status_color = "✕ İptal"
        else:
            status_color = status

        # Type badge
        tx_type = tx.get("type", "")
        type_label = "Gelir" if tx_type == "income" else ("Gider" if tx_type == "expense" else tx_type)

        # Amount
        amount = tx.get("amount", 0)
        amount_str = _money(amount) if tx_type == "income" else f"({_money(amount)})"

        values = [
            tx_date.strftime("%d.%m.%Y") if tx_date else "",
            tx.get("description", "") or tx.get("note", "") or "",
            tx.get("category", "") or "",
            type_label,
            amount_str,
            status_color,
        ]
        for j, v in enumerate(values):
            cell = ws.cell(row=row, column=j + 1, value=v)
            cell.font = _body_font()
            cell.border = _border
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Summary row
    total_income = sum(tx.get("amount", 0) for tx in transactions if tx.get("type") == "income")
    total_expense = sum(tx.get("amount", 0) for tx in transactions if tx.get("type") == "expense")
    summary_row = len(transactions) + 4
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    s_cell = ws.cell(row=summary_row, column=1, value="TOPLAM")
    s_cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    s_cell.fill = SUBHEADER_FILL
    s_cell.alignment = Alignment(horizontal="right")
    s_cell.border = _border

    income_cell = ws.cell(row=summary_row, column=5, value=_money(total_income))
    income_cell.font = Font(name="Calibri", size=11, bold=True, color="6B8E6B")
    income_cell.fill = SUBHEADER_FILL
    income_cell.border = _border

    expense_cell = ws.cell(row=summary_row, column=6, value=f"({_money(total_expense)})")
    expense_cell.font = Font(name="Calibri", size=11, bold=True, color="AD7B7B")
    expense_cell.fill = SUBHEADER_FILL
    expense_cell.border = _border

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Analytics / BI ────────────────────────────────────────────────────────

def export_analytics_to_excel(dashboard_data: Dict[str, Any], agency_name: str = "") -> bytes:
    """
    BI Analytics verilerini Excel'e döker — PRD §4.1.10.
    Çoklu sheet: Özet | Doluluk | Kiracı Performansı | Finans
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Özet ──────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Özet")
    _build_summary_sheet(ws_summary, dashboard_data, agency_name)

    # ── Sheet 2: Doluluk ───────────────────────────────────────────────────
    ws_occ = wb.create_sheet("Doluluk")
    _build_occupancy_sheet(ws_occ, dashboard_data)

    # ── Sheet 3: Kiracı Performans ─────────────────────────────────────────
    ws_tenants = wb.create_sheet("Kiracı Performansı")
    _build_tenant_tabsheet(ws_tenants, dashboard_data)

    # ── Sheet 4: Finans ────────────────────────────────────────────────────
    ws_fin = wb.create_sheet("Finans")
    _build_financial_tabsheet(ws_fin, dashboard_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_summary_sheet(ws, data: Dict[str, Any], agency_name: str):
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"{agency_name} — BI Özet Raporu"
    t.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor="1A1A1A")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    d = ws["A2"]
    d.value = f"Rapor Tarihi: {date.today().strftime('%d.%m.%Y')}"
    d.font = _body_font(size=9, color="888888")
    d.fill = PatternFill("solid", fgColor="1A1A1A")

    kpis = data.get("kpis", {})
    headers = ["Metrik", "Değer"]
    _set_header(ws, 3, headers)

    rows = [
        ("Toplam Mülk", kpis.get("total_properties", "")),
        ("Toplam Birim", kpis.get("total_units", "")),
        ("Dolu Birim", kpis.get("occupied_units", "")),
        ("Boş Birim", kpis.get("vacant_units", "")),
        ("Doluluk Oranı", f"%{kpis.get('occupancy_rate', 0):.1f}"),
        ("Aktif Kiracı", kpis.get("active_tenants", "")),
        ("Bu Ay Tahsilat", _money(kpis.get("this_month_collected", 0))),
        ("Bekleyen", _money(kpis.get("pending_this_month", 0))),
        ("Gecikmiş", _money(kpis.get("overdue_amount", 0))),
    ]
    for i, (label, value) in enumerate(rows):
        row = i + 4
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = _body_font()
        c1.border = _border
        if i % 2 == 1:
            c1.fill = ALT_ROW_FILL
        c2 = ws.cell(row=row, column=2, value=str(value))
        c2.font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
        c2.border = _border
        if i % 2 == 1:
            c2.fill = ALT_ROW_FILL

    _auto_width(ws)


def _build_occupancy_sheet(ws, data: Dict[str, Any]):
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "Doluluk Analizi"
    t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor="1A1A1A")
    ws.row_dimensions[1].height = 24

    trend = data.get("occupancy_trend", [])
    if trend:
        _set_header(ws, 2, ["Ay", "Doluluk Oranı (%)"])
        for i, item in enumerate(trend):
            row = i + 3
            for j, key in enumerate(["month", "occupancy_rate"]):
                val = item.get(key, "")
                cell = ws.cell(row=row, column=j + 1, value=(
                    f"%{val:.1f}" if key == "occupancy_rate" else val
                ))
                cell.font = _body_font()
                cell.border = _border
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL
    else:
        ws["A2"] = "Veri mevcut değil"

    _auto_width(ws)


def _build_tenant_tabsheet(ws, data: Dict[str, Any]):
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Kiracı Devir Analizi"
    t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor="1A1A1A")
    ws.row_dimensions[1].height = 24

    headers = ["Ay", "Yeni Kiracı", "Ayrılan Kiracı", "Aktif Kiracı"]
    _set_header(ws, 2, headers)

    churn = data.get("tenant_churn", {})
    active = churn.get("total_active_tenants", 0)
    flow = churn.get("monthly_flow", [])

    # Global active count (cumulative)
    cumulative = active
    rows_data = []
    for f in reversed(flow):
        cumulative += f.get("departed_tenants", 0) - f.get("new_tenants", 0)
        rows_data.insert(0, {
            "month": f.get("month", ""),
            "new": f.get("new_tenants", 0),
            "departed": f.get("departed_tenants", 0),
            "active": cumulative,
        })

    for i, row_data in enumerate(rows_data):
        row = i + 3
        values = [row_data["month"], row_data["new"], row_data["departed"], row_data["active"]]
        for j, v in enumerate(values):
            cell = ws.cell(row=row, column=j + 1, value=v)
            cell.font = _body_font()
            cell.border = _border
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    _auto_width(ws)


def _build_financial_tabsheet(ws, data: Dict[str, Any]):
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "Yıllık Finansal Özet"
    t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor="1A1A1A")
    ws.row_dimensions[1].height = 24

    fin = data.get("financial_annual", {})
    headers = ["Ay", "Toplam Gelir", "Tahsilat", "Gider", "Net"]
    _set_header(ws, 2, headers)

    monthly = fin.get("monthly_breakdown", [])
    for i, m in enumerate(monthly):
        row = i + 3
        values = [
            m.get("month", ""),
            _money(m.get("total_income", 0)),
            _money(m.get("total_income", 0)),  # collected ≈ total_income
            _money(m.get("total_expense", 0)),
            _money(m.get("net_balance", 0)),
        ]
        for j, v in enumerate(values):
            cell = ws.cell(row=row, column=j + 1, value=v)
            cell.font = _body_font()
            cell.border = _border
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Annual totals
    total_row = len(monthly) + 3
    ws.cell(row=total_row, column=1, value="TOPLAM").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=total_row, column=1).border = _border
    for j, key in enumerate(["total_income", "total_income", "total_expense", "net_balance"]):
        val = sum(m.get(key, 0) for m in monthly)
        cell = ws.cell(row=total_row, column=j + 2, value=_money(val))
        cell.font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
        cell.fill = SUBHEADER_FILL
        cell.border = _border

    _auto_width(ws)
